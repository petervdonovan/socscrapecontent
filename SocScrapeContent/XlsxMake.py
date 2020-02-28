from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill


class XlsxMake:
    '''
    Class for creating the spreadsheet using 2D lists of data about the school site.
    '''
    def __init__(self, pathwaysData, clubsData):
        '''
        Initialize with 2D arrays of the data needed; create workbook with the 3 sheets
        into which the data will be transferred.
        '''
        self.pathwaysData = pathwaysData
        self.clubsData = clubsData
        # Create a workbook with the 3 appropriate sheets
        self.wb = Workbook()
        self.pathwaysWs = self.wb.active
        self.pathwaysWs.title = 'Pathways'
        self.clubsWs = self.wb.create_sheet('Clubs')
        self.athleticsWs = self.wb.create_sheet('Athletics')

    
    def fill(self):
        '''
        Transfer all data from the 2D lists to the spreadsheet
        '''
        for row in self.pathwaysData:
            self.pathwaysWs.append(row)
        for row in self.clubsData:
            self.clubsWs.append(row)
    
    def save(self, filename):
        '''
        Save the workbook as an XLSX file in the sheets folder in the project folder
        '''
        try:
            self.wb.save(filename = 'sheets/' + filename)
        except(IOError):
            print('Unable to save. Is the file you intend to write to already open in another program?')
            input('Press enter when you have closed the file.')
            self.save(filename)
    
    def make(self, sectionWidth, sectionHeight):
        '''
        Create workbook.
        '''
        self.fill()
        self.formatPathwayWs(sectionWidth, sectionHeight)
        self.formatClubsWs()
        self.formatAthleticsWs()
    
    def formatPathwayWs(self, sectionWidth, sectionHeight):
        '''
        Formatting (borders, bold text, size) for pathway sheet in the workbook
        '''
        for topRow in range(1, len(self.pathwaysData), sectionHeight):
            # Define different possible borders
            thin = Side(border_style="thin", color="000000")
            thick = Side(border_style="thick", color="000000")
            none = Side(border_style=None, color='FF000000')
            # Iterate over rows based on the sectionHeight found in the cell at the top of the template XLSX.
            for row in range(topRow, topRow + sectionHeight):
                for col in range(1, sectionWidth + 1):
                    # Default values
                    size = 12
                    bold = False
                    cell = self.pathwaysWs.cell(row, col)
                    # Iff the cell contains information, give it a border
                    if (cell.value):
                        left = right = top = bottom = thin
                    else:
                        left = right = top = bottom = none
                    # The outside borders of the entire section are thick
                    # Left border
                    if col == 0:
                        left = thick
                    # Right border
                    if col == sectionWidth:
                        right = thick
                    # Top border
                    if row == topRow:
                        top = thick
                    # Bottom border
                    if row == topRow + sectionHeight - 1:
                        bottom = thick
                    #Style headers
                    cell = self.pathwaysWs.cell(row, col)
                    if (cell.value and (row - topRow < 2 or not self.pathwaysWs.cell(row - 1, col).value)):
                        bold = True
                    #handle exceptions on which headers should be bold
                    if (row - topRow == 16 and col == 6): bold = True
                    if (row - topRow == 17 and col == 4): bold = False

                    #Style top row
                    if row == topRow and col == 1:
                        left = top = right = left = thick
                        size = 18
                    # Apply the border & font styles chosen using the above logic
                    cell.border = Border(left = left, right = right, top = top, bottom = bottom)
                    cell.font = Font(size = size, bold = bold)
                    # Merge cells in the top row
                    self.pathwaysWs.merge_cells(start_row = topRow, start_column = 1, end_row = topRow, end_column = sectionWidth)
    def formatClubsWs(self):
        '''
        Add formatting (borders, colors) to existing Clubs worksheet
        '''
        # Define the types of borders a cell can have
        thin = Side(border_style="thin", color="666666")
        thick = Side(border_style="thick", color="000000")
        # The final column is column 8 (counting from 1)
        maxCol = 8;
        # Iterate over all rows (or the first 10,000 rows, which hopefully will never happen)
        for row in self.clubsWs.iter_rows(min_row=1, max_row=10000, min_col=1, max_col=maxCol):
            color = '000000'
            cells = [cell for cell in row]
            # Break out of the loop if row is blank (so that it does not have to fill in blank rows until there are 10,000 rows)
            if not cells[0].value:
                break;
            # the 3rd cell in every row states the category (Arts, Cultural, STEM, Volunteering and Advocacy, etc.)
            category = cells[3].value.lower()
            # Color the rows based on the category
            if(category == 'arts'):
                color = 'ff8888'
            elif(category == 'cultural'):
                color = '88ff88'
            elif(category == 'volunteering and advocacy'):
                color = '8888ff'
            elif(category == 'social'):
                color = '88ffff'
            elif(category == 'stem'):
                color = 'ff88ff'
            elif(category == 'other'):
                color = 'ffff88'
            # Apply formatting to all the cells
            for cell in cells:
                # Apply chosen color
                cell.fill = PatternFill("solid", fgColor=color)
                # default: all borders are thin
                left = right = top = bottom = thin
                # Rightmost border of the content in the sheet is thick
                if (cell.column == maxCol):
                    right = thick
                # Black cells are the only ones that can't have black text. They have white text
                if(color == '000000'):
                    cell.font = Font(color="ffffff")
                # Apply chosen border
                cell.border = Border(left = left, right = right, top = top, bottom = bottom)
    
    def formatAthleticsWs(self):
        '''
        The Athletics worksheet will just be blank so that the user can input the values manually, so 
        this function creates the headers and formats them
        '''
        # This is row 1 of the worksheet
        head = ["Sport", "Gender (Boys, Girls, Coed)",	"Months in season",	"Month in off-season training", "CIF -- years gone to playoffs (since 2000)", "Moore League -- years won (since 2000)", "Other tournaments/awards"]
        col = 1
        # Input the values of everything in row 1 and style it
        for item in head:
            # Copy the value into the athletics worksheet
            cell = self.athleticsWs.cell(1, col)
            cell.value = item
            cell.fill = PatternFill("solid", fgColor='000000')
            cell.font = Font(color="ffffff")
            col += 1